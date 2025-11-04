"""Portfolio export utilities for PDF and Excel."""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def _format_currency(value: Any) -> str:
    """Format currency for display."""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return "N/A"
    magnitude = abs(number)
    if magnitude >= 1_000_000_000:
        return f"${number / 1_000_000_000:.2f}B"
    if magnitude >= 1_000_000:
        return f"${number / 1_000_000:.2f}M"
    if magnitude >= 1_000:
        return f"${number / 1_000:.2f}K"
    return f"${number:,.2f}"


def _format_percent(value: Any) -> str:
    """Format percentage."""
    try:
        number = float(value)
        return f"{number * 100:.2f}%"
    except (TypeError, ValueError):
        return "N/A"


def _sanitize_text_for_pdf(text: str) -> str:
    """Remove or replace Unicode characters that FPDF cannot handle."""
    if not text:
        return ""
    
    replacements = {
        '\u2022': '-',      # Bullet point → hyphen
        '\u2013': '-',      # En dash → hyphen
        '\u2014': '--',     # Em dash → double hyphen
        '\u2018': "'",      # Left single quote → apostrophe
        '\u2019': "'",      # Right single quote → apostrophe
        '\u201c': '"',      # Left double quote → quote
        '\u201d': '"',      # Right double quote → quote
        '\u2026': '...',    # Ellipsis → three dots
        '\x07': ' ',        # Bell character → space
        '\u00a0': ' ',      # Non-breaking space → space
    }
    
    for unicode_char, ascii_char in replacements.items():
        text = text.replace(unicode_char, ascii_char)
    
    # Remove any remaining non-Latin-1 characters
    try:
        text.encode('latin-1')
    except UnicodeEncodeError:
        text = text.encode('ascii', errors='ignore').decode('ascii')
    
    return text


def build_portfolio_pdf(
    portfolio_id: str,
    portfolio_name: str,
    holdings: List[Dict[str, Any]],
    summary: Dict[str, Any],
    exposure: Dict[str, Any],
    attribution: Optional[Dict[str, Any]] = None,
    optimization: Optional[Dict[str, Any]] = None,
) -> bytes:
    """Generate 1-2 page portfolio executive summary PDF."""
    
    as_of = datetime.utcnow().strftime("%B %d, %Y")
    
    # Initialize PDF
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(left=20, top=20, right=20)
    
    # Color scheme (professional blues and grays)
    COLOR_PRIMARY = (11, 33, 74)      # Navy blue
    COLOR_SECONDARY = (30, 90, 168)  # Light blue
    COLOR_ACCENT = (50, 50, 50)       # Dark gray
    COLOR_LIGHT_GRAY = (240, 240, 240)
    COLOR_TEXT = (40, 40, 40)
    
    # ========== COVER PAGE ==========
    pdf.add_page()
    
    # Header bar
    pdf.set_fill_color(*COLOR_PRIMARY)
    pdf.rect(0, 0, 210, 40, 'F')
    
    # Portfolio name
    pdf.set_y(12)
    pdf.set_font("Helvetica", "B", 26)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, _sanitize_text_for_pdf(portfolio_name), align="C", ln=1)
    
    # Portfolio ID
    pdf.set_font("Helvetica", "", 16)
    pdf.set_text_color(220, 220, 255)
    pdf.cell(0, 8, f"({portfolio_id})", align="C", ln=1)
    
    # Report title
    pdf.set_y(60)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*COLOR_ACCENT)
    pdf.cell(0, 10, "Portfolio Analysis Report", align="C", ln=1)
    
    # Decorative line
    pdf.set_y(75)
    pdf.set_draw_color(*COLOR_SECONDARY)
    pdf.set_line_width(0.5)
    pdf.line(70, 75, 140, 75)
    
    # Executive Summary Box
    pdf.set_y(90)
    pdf.set_fill_color(*COLOR_LIGHT_GRAY)
    pdf.set_draw_color(*COLOR_SECONDARY)
    pdf.set_line_width(0.3)
    pdf.rect(30, 90, 150, 50, 'FD')
    
    pdf.set_y(95)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(0, 8, "Executive Summary", align="C", ln=1)
    
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(*COLOR_TEXT)
    
    # Key metrics
    pdf.set_x(40)
    pdf.cell(60, 8, "Total Value:", align="L")
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, _format_currency(summary.get('total_market_value', 0)), align="L", ln=1)
    
    pdf.set_font("Helvetica", "", 11)
    pdf.set_x(40)
    pdf.cell(60, 8, "Holdings:", align="L")
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, f"{summary.get('num_holdings', 0)} positions", align="L", ln=1)
    
    pdf.set_font("Helvetica", "", 11)
    pdf.set_x(40)
    pdf.cell(60, 8, "Sectors:", align="L")
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, f"{summary.get('num_sectors', 0)} sectors", align="L", ln=1)
    
    pdf.set_font("Helvetica", "", 11)
    pdf.set_x(40)
    pdf.cell(60, 8, "Top 10 Concentration:", align="L")
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, _format_percent(summary.get('top_10_concentration', 0)), align="L", ln=1)
    
    # Footer
    pdf.set_y(260)
    pdf.set_font("Helvetica", "I", 10)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Report Date: {as_of}", align="C", ln=1)
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(0, 6, "Generated by BenchmarkOS | Team 2", align="C", ln=1)
    
    # ========== DETAILED ANALYSIS PAGE ==========
    pdf.add_page()
    
    # Header
    pdf.set_fill_color(*COLOR_PRIMARY)
    pdf.rect(0, 0, 210, 30, 'F')
    pdf.set_y(8)
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 8, _sanitize_text_for_pdf(portfolio_name), align="L", ln=1)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(220, 220, 255)
    pdf.cell(0, 6, f"{portfolio_id} | {as_of.upper()}", align="L", ln=1)
    
    pdf.set_y(40)
    
    # Portfolio Overview Section
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(0, 8, "Portfolio Overview", ln=1)
    
    pdf.set_draw_color(*COLOR_SECONDARY)
    pdf.set_line_width(0.5)
    pdf.line(20, 50, 190, 50)
    pdf.ln(5)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*COLOR_TEXT)
    
    # Key metrics
    metrics_text = [
        f"Total Market Value: {_format_currency(summary.get('total_market_value', 0))}",
        f"Number of Holdings: {summary.get('num_holdings', 0)}",
        f"Number of Sectors: {summary.get('num_sectors', 0)}",
        f"Weighted Average P/E: {summary.get('weighted_avg_pe', 0):.1f}×" if summary.get('weighted_avg_pe') else "P/E: N/A",
        f"Weighted Average Dividend Yield: {_format_percent(summary.get('weighted_avg_dividend_yield', 0))}",
        f"Top 10 Concentration: {_format_percent(summary.get('top_10_concentration', 0))}",
    ]
    
    for metric in metrics_text:
        pdf.cell(0, 6, _sanitize_text_for_pdf(f"• {metric}"), ln=1)
    
    pdf.ln(5)
    
    # Top Holdings Section
    pdf.set_font("Helvetica", "B", 14)
    pdf.set_text_color(*COLOR_PRIMARY)
    pdf.cell(0, 8, "Top 10 Holdings", ln=1)
    
    pdf.set_draw_color(*COLOR_SECONDARY)
    pdf.set_line_width(0.5)
    pdf.line(20, pdf.get_y() + 2, 190, pdf.get_y() + 2)
    pdf.ln(5)
    
    # Table header
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(255, 255, 255)
    pdf.set_fill_color(*COLOR_PRIMARY)
    pdf.set_draw_color(255, 255, 255)
    
    header_widths = [40, 30, 40, 30, 50]
    headers = ["Ticker", "Weight", "Sector", "P/E", "Market Value"]
    
    x_start = 20
    for i, (header, width) in enumerate(zip(headers, header_widths)):
        pdf.set_xy(x_start + sum(header_widths[:i]), pdf.get_y())
        pdf.cell(width, 8, _sanitize_text_for_pdf(header), border=1, align="C", fill=True)
    pdf.ln()
    
    # Table rows
    top_holdings = sorted(holdings, key=lambda h: h.get('weight', 0), reverse=True)[:10]
    
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(*COLOR_TEXT)
    
    for holding in top_holdings:
        pdf.set_fill_color(*COLOR_LIGHT_GRAY)
        
        ticker = _sanitize_text_for_pdf(holding.get('ticker', 'N/A'))
        weight = _format_percent(holding.get('weight', 0))
        sector = _sanitize_text_for_pdf(holding.get('sector', 'N/A'))
        pe = f"{holding.get('pe_ratio', 0):.1f}×" if holding.get('pe_ratio') else "N/A"
        market_value = _format_currency(holding.get('market_value', 0))
        
        row_data = [ticker, weight, sector, pe, market_value]
        
        for i, (data, width) in enumerate(zip(row_data, header_widths)):
            pdf.set_xy(x_start + sum(header_widths[:i]), pdf.get_y())
            pdf.cell(width, 6, _sanitize_text_for_pdf(data), border=1, align="L", fill=True)
        pdf.ln()
    
    pdf.ln(5)
    
    # Sector Exposures Section
    sector_exposure = exposure.get('sector_exposure', {})
    if sector_exposure:
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(*COLOR_PRIMARY)
        pdf.cell(0, 8, "Sector Exposures", ln=1)
        
        pdf.set_draw_color(*COLOR_SECONDARY)
        pdf.set_line_width(0.5)
        pdf.line(20, pdf.get_y() + 2, 190, pdf.get_y() + 2)
        pdf.ln(5)
        
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*COLOR_TEXT)
        
        sorted_sectors = sorted(sector_exposure.items(), key=lambda x: x[1], reverse=True)
        for sector, weight in sorted_sectors[:10]:  # Top 10 sectors
            pdf.cell(0, 6, _sanitize_text_for_pdf(f"• {sector}: {_format_percent(weight)}"), ln=1)
    
    # Performance Attribution (if available)
    if attribution:
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 14)
        pdf.set_text_color(*COLOR_PRIMARY)
        pdf.cell(0, 8, "Performance Attribution", ln=1)
        
        pdf.set_draw_color(*COLOR_SECONDARY)
        pdf.set_line_width(0.5)
        pdf.line(20, pdf.get_y() + 2, 190, pdf.get_y() + 2)
        pdf.ln(5)
        
        pdf.set_font("Helvetica", "", 10)
        pdf.set_text_color(*COLOR_TEXT)
        
        attr_text = [
            f"Total Active Return: {_format_percent(attribution.get('total_active_return', 0))}",
            f"Allocation Effect: {_format_percent(sum(attribution.get('allocation_effect', {}).values()))}",
            f"Selection Effect: {_format_percent(sum(attribution.get('selection_effect', {}).values()))}",
            f"Interaction Effect: {_format_percent(sum(attribution.get('interaction_effect', {}).values()))}",
        ]
        
        for text in attr_text:
            pdf.cell(0, 6, _sanitize_text_for_pdf(f"• {text}"), ln=1)
    
    # Footer
    pdf.set_y(270)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 5, f"Analysis Date: {as_of} | Portfolio ID: {portfolio_id}", align="C", ln=1)
    pdf.cell(0, 5, "Generated by BenchmarkOS Intelligence Platform | Team 2", align="C", ln=1)
    
    # Save to bytes
    return pdf.output(dest='S').encode('latin-1')


def build_portfolio_excel(
    portfolio_id: str,
    portfolio_name: str,
    holdings: List[Dict[str, Any]],
    summary: Dict[str, Any],
    exposure: Dict[str, Any],
    attribution: Optional[Dict[str, Any]] = None,
    optimization: Optional[Dict[str, Any]] = None,
    scenarios: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """Generate multi-tab Excel workbook with portfolio analysis."""
    
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="0B214A", end_color="0B214A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== HOLDINGS TAB ==========
    ws_holdings = wb.active
    ws_holdings.title = "Holdings"
    
    # Header
    headers = ["Ticker", "Weight", "Shares", "Price", "Market Value", "Sector", "P/E Ratio", "Dividend Yield", "ROE", "ROIC"]
    for col, header in enumerate(headers, 1):
        cell = ws_holdings.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    
    # Data rows
    sorted_holdings = sorted(holdings, key=lambda h: h.get('weight', 0), reverse=True)
    
    for row_idx, holding in enumerate(sorted_holdings, start=2):
        ws_holdings.cell(row=row_idx, column=1, value=holding.get('ticker', ''))
        ws_holdings.cell(row=row_idx, column=2, value=holding.get('weight', 0))
        ws_holdings.cell(row=row_idx, column=3, value=holding.get('shares', 0))
        ws_holdings.cell(row=row_idx, column=4, value=holding.get('price', 0))
        ws_holdings.cell(row=row_idx, column=5, value=holding.get('market_value', 0))
        ws_holdings.cell(row=row_idx, column=6, value=holding.get('sector', ''))
        ws_holdings.cell(row=row_idx, column=7, value=holding.get('pe_ratio'))
        ws_holdings.cell(row=row_idx, column=8, value=holding.get('dividend_yield', 0))
        ws_holdings.cell(row=row_idx, column=9, value=holding.get('roe'))
        ws_holdings.cell(row=row_idx, column=10, value=holding.get('roic'))
        
        # Apply border and alignment
        for col in range(1, 11):
            cell = ws_holdings.cell(row=row_idx, column=col)
            cell.border = border
            if col in [2, 3, 4, 5, 7, 8, 9, 10]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col in range(1, 11):
        ws_holdings.column_dimensions[chr(64 + col)].width = 15
    
    # ========== EXPOSURES TAB ==========
    ws_exposures = wb.create_sheet("Exposures")
    
    # Sector Exposures
    ws_exposures.cell(row=1, column=1, value="Sector Exposures").font = Font(bold=True, size=12)
    ws_exposures.cell(row=2, column=1, value="Sector").fill = header_fill
    ws_exposures.cell(row=2, column=1, value="Sector").font = header_font
    ws_exposures.cell(row=2, column=2, value="Weight").fill = header_fill
    ws_exposures.cell(row=2, column=2, value="Weight").font = header_font
    
    sector_exposure = exposure.get('sector_exposure', {})
    sorted_sectors = sorted(sector_exposure.items(), key=lambda x: x[1], reverse=True)
    
    for row_idx, (sector, weight) in enumerate(sorted_sectors, start=3):
        ws_exposures.cell(row=row_idx, column=1, value=sector)
        ws_exposures.cell(row=row_idx, column=2, value=weight)
        ws_exposures.cell(row=row_idx, column=2).number_format = '0.00%'
    
    # Factor Exposures
    row_start = len(sorted_sectors) + 5
    ws_exposures.cell(row=row_start, column=1, value="Factor Exposures").font = Font(bold=True, size=12)
    ws_exposures.cell(row=row_start + 1, column=1, value="Factor").fill = header_fill
    ws_exposures.cell(row=row_start + 1, column=1, value="Factor").font = header_font
    ws_exposures.cell(row=row_start + 1, column=2, value="Exposure").fill = header_fill
    ws_exposures.cell(row=row_start + 1, column=2, value="Exposure").font = header_font
    
    factor_exposure = exposure.get('factor_exposure', {})
    sorted_factors = sorted(factor_exposure.items(), key=lambda x: abs(x[1]), reverse=True)
    
    for row_idx, (factor, exposure_value) in enumerate(sorted_factors, start=row_start + 2):
        ws_exposures.cell(row=row_idx, column=1, value=factor.title())
        ws_exposures.cell(row=row_idx, column=2, value=exposure_value)
    
    # Auto-adjust column widths
    ws_exposures.column_dimensions['A'].width = 25
    ws_exposures.column_dimensions['B'].width = 15
    
    # ========== SUMMARY TAB ==========
    ws_summary = wb.create_sheet("Summary")
    
    ws_summary.cell(row=1, column=1, value="Portfolio Summary").font = Font(bold=True, size=14)
    
    summary_data = [
        ["Portfolio ID", portfolio_id],
        ["Portfolio Name", portfolio_name],
        ["Total Market Value", summary.get('total_market_value', 0)],
        ["Number of Holdings", summary.get('num_holdings', 0)],
        ["Number of Sectors", summary.get('num_sectors', 0)],
        ["Weighted Avg P/E", summary.get('weighted_avg_pe')],
        ["Weighted Avg Dividend Yield", summary.get('weighted_avg_dividend_yield', 0)],
        ["Top 10 Concentration", summary.get('top_10_concentration', 0)],
    ]
    
    ws_summary.cell(row=3, column=1, value="Metric").fill = header_fill
    ws_summary.cell(row=3, column=1, value="Metric").font = header_font
    ws_summary.cell(row=3, column=2, value="Value").fill = header_fill
    ws_summary.cell(row=3, column=2, value="Value").font = header_font
    
    for row_idx, (label, value) in enumerate(summary_data, start=4):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=value)
    
    # Sector Breakdown
    row_start = len(summary_data) + 6
    ws_summary.cell(row=row_start, column=1, value="Sector Breakdown").font = Font(bold=True, size=12)
    ws_summary.cell(row=row_start + 1, column=1, value="Sector").fill = header_fill
    ws_summary.cell(row=row_start + 1, column=1, value="Sector").font = header_font
    ws_summary.cell(row=row_start + 1, column=2, value="Weight").fill = header_fill
    ws_summary.cell(row=row_start + 1, column=2, value="Weight").font = header_font
    
    sector_breakdown = summary.get('sector_breakdown', {})
    sorted_sectors = sorted(sector_breakdown.items(), key=lambda x: x[1], reverse=True)
    
    for row_idx, (sector, weight) in enumerate(sorted_sectors, start=row_start + 2):
        ws_summary.cell(row=row_idx, column=1, value=sector)
        ws_summary.cell(row=row_idx, column=2, value=weight)
        ws_summary.cell(row=row_idx, column=2).number_format = '0.00%'
    
    # Auto-adjust column widths
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 20
    
    # ========== ATTRIBUTION TAB ==========
    if attribution:
        ws_attribution = wb.create_sheet("Attribution")
        
        ws_attribution.cell(row=1, column=1, value="Performance Attribution").font = Font(bold=True, size=14)
        
        # Summary metrics
        attr_summary = [
            ["Total Active Return", attribution.get('total_active_return', 0)],
            ["Allocation Effect", sum(attribution.get('allocation_effect', {}).values())],
            ["Selection Effect", sum(attribution.get('selection_effect', {}).values())],
            ["Interaction Effect", sum(attribution.get('interaction_effect', {}).values())],
        ]
        
        ws_attribution.cell(row=3, column=1, value="Metric").fill = header_fill
        ws_attribution.cell(row=3, column=1, value="Metric").font = header_font
        ws_attribution.cell(row=3, column=2, value="Value").fill = header_fill
        ws_attribution.cell(row=3, column=2, value="Value").font = header_font
        
        for row_idx, (label, value) in enumerate(attr_summary, start=4):
            ws_attribution.cell(row=row_idx, column=1, value=label)
            ws_attribution.cell(row=row_idx, column=2, value=value)
            ws_attribution.cell(row=row_idx, column=2).number_format = '0.00%'
        
        # Top Contributors
        row_start = len(attr_summary) + 6
        ws_attribution.cell(row=row_start, column=1, value="Top Contributors").font = Font(bold=True, size=12)
        
        top_contributors = attribution.get('top_contributors', [])[:10]
        if top_contributors:
            headers = ["Ticker", "Contribution", "Sector"]
            for col, header in enumerate(headers, 1):
                cell = ws_attribution.cell(row=row_start + 1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, contrib in enumerate(top_contributors, start=row_start + 2):
                ticker = contrib.get('ticker', '')
                holding = next((h for h in holdings if h.get('ticker') == ticker), {})
                ws_attribution.cell(row=row_idx, column=1, value=ticker)
                ws_attribution.cell(row=row_idx, column=2, value=contrib.get('contribution', 0))
                ws_attribution.cell(row=row_idx, column=2).number_format = '0.00%'
                ws_attribution.cell(row=row_idx, column=3, value=holding.get('sector', ''))
        
        # Auto-adjust column widths
        ws_attribution.column_dimensions['A'].width = 20
        ws_attribution.column_dimensions['B'].width = 15
        ws_attribution.column_dimensions['C'].width = 20
    
    # ========== OPTIMIZATION TAB ==========
    if optimization and optimization.get('proposed_trades'):
        ws_optimization = wb.create_sheet("Optimization")
        
        ws_optimization.cell(row=1, column=1, value="Optimization Recommendations").font = Font(bold=True, size=14)
        
        headers = ["Ticker", "Action", "From Weight", "To Weight", "Shares"]
        for col, header in enumerate(headers, 1):
            cell = ws_optimization.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        trades = optimization.get('proposed_trades', [])
        for row_idx, trade in enumerate(trades, start=4):
            ws_optimization.cell(row=row_idx, column=1, value=trade.get('ticker', ''))
            ws_optimization.cell(row=row_idx, column=2, value=trade.get('action', 'hold').upper())
            ws_optimization.cell(row=row_idx, column=3, value=trade.get('from_weight', 0))
            ws_optimization.cell(row=row_idx, column=3).number_format = '0.00%'
            ws_optimization.cell(row=row_idx, column=4, value=trade.get('to_weight', 0))
            ws_optimization.cell(row=row_idx, column=4).number_format = '0.00%'
            ws_optimization.cell(row=row_idx, column=5, value=trade.get('shares', 0))
        
        # Auto-adjust column widths
        for col in range(1, 6):
            ws_optimization.column_dimensions[chr(64 + col)].width = 15
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


